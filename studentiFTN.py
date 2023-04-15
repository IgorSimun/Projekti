#All the necesarry imports.
import argparse
import openpyxl
from openpyxl.chart import Reference, BarChart, LineChart
from openpyxl.styles import Font, Alignment
import datetime

#Base dictionary for filling out of the excel file.
dictUcenika={

    'Milos': [2018, "AE258", 9.54, 10],
    'Marko': [2017, "AE451", 8.27, 10],
    'Darko': [2016, "AE625", 6.48, 8]
}

#Defining of lists whose values will be used in excel file.
imeUcenika = list(dictUcenika.keys())
vrednosti = list(dictUcenika.values())
godinaUpisa = []
brojIndeksa= []
prosecnaOcena= []
najvisaOcena = []

#Adding values of dictionary and lists previously defined ,to the excel file.
for x in range(len(imeUcenika)):
    godinaUpisa.append(vrednosti[x][0])
    brojIndeksa.append(vrednosti[x][1])
    prosecnaOcena.append(vrednosti[x][2])
    najvisaOcena.append(vrednosti[x][3])

#Opening of a workbook in excel file and defining of dimensions of rows and columns, also manual adding of necessary values.
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Studenti"

ws.column_dimensions['A'].width = 13
ws.column_dimensions['B'].width = 13
ws.column_dimensions['C'].width = 13
ws.column_dimensions['D'].width = 13
ws.column_dimensions['E'].width = 15
ws.column_dimensions['F'].width = 15

font = Font(bold=True)

ws['A1'].value = 'Redni broj'
ws['A1'].font = font
ws['B1'].value = 'Ime ucenika'
ws['B1'].font = font
ws['C1'].value = 'Godina upisa'
ws['C1'].font = font
ws['D1'].value = 'Broj indeksa'
ws['D1'].font = font
ws['E1'].value = 'Prosecna ocena'
ws['E1'].font = font
ws['F1'].value = 'Najvisa ocena'
ws['F1'].font = font
    
#Manual adding of new students and their infos.
def addStudents():
    if students:
        while True:
            unos = input("Would you like to add more students? y/n: ").lower()
            if unos == 'y':
                imeUcenika.append(input("Unesite ime ucenika: "))
                godinaUpisa.append(int(input("Unesite godinu upisa: ")))
                brojIndeksa.append(input("Unesite broj indeksa: "))
                prosecnaOcena.append(float(input("Unesite prosecnu ocenu: ")))
                najvisaOcena.append(int(input("Unesite najvisu ocenu: ")))
            elif unos == 'n':
                break
            else:
                print("Pogresan unos, pokusajte ponovo")
                continue

    #Adding serial numbers and values in excel file, also defining of an alignment.
    for i in range(len(imeUcenika)):
        row_number = i + 2
        ws['A' + str(row_number)].value = i + 1
        ws['A' + str(row_number)].font = Font(bold=True)
        ws['A' + str(row_number)].alignment = Alignment(horizontal='right')
        ws['B' + str(row_number)].value = imeUcenika[i]
        ws['B' + str(row_number)].alignment = Alignment(horizontal='right')
        ws['C' + str(row_number)].value = godinaUpisa[i]
        ws['C' + str(row_number)].alignment = Alignment(horizontal='right')
        ws['D' + str(row_number)].value = brojIndeksa[i]
        ws['D' + str(row_number)].alignment = Alignment(horizontal='right')
        ws['E' + str(row_number)].value = str(prosecnaOcena[i]).replace('.',',')
        ws['E' + str(row_number)].alignment = Alignment(horizontal='right')
        ws['F' + str(row_number)].value = str(najvisaOcena[i]).replace('.',',')
        ws['F' + str(row_number)].alignment = Alignment(horizontal='right')

#Adding a new row "Proseci", and defining alignment in exact rows.
def averageAmount():
    today = datetime.date.today()
    yearNow = today.year
    for i in range(len(imeUcenika)):
        ws['A' + str(len(imeUcenika)+ 2)].value = 'Proseci: '
        ws['A' + str(len(imeUcenika)+ 2)].font = Font(bold=True)
        ws['B' + str(len(imeUcenika)+ 2)].value = int(len(imeUcenika))
        ws['B' + str(len(imeUcenika)+ 2)].alignment = Alignment(horizontal='right')
        ws['C' + str(len(imeUcenika)+ 2)].value = round(yearNow - sum(godinaUpisa)/len(imeUcenika))
        ws['C' + str(len(imeUcenika)+ 2)].alignment = Alignment(horizontal='right')
        ws['E' + str(len(imeUcenika)+ 2)].value = str(float(round(sum(prosecnaOcena)/len(imeUcenika),2))).replace('.',',')
        ws['E' + str(len(imeUcenika)+ 2)].alignment = Alignment(horizontal='right')
        ws['F' + str(len(imeUcenika)+ 2)].value = str(float(round(sum(najvisaOcena)/len(imeUcenika),2))).replace('.',',')
        ws['F' + str(len(imeUcenika)+ 2)].alignment = Alignment(horizontal='right')

# Drawing a chart.
def createChart():
    ws1 = wb.create_sheet("Grafik")
    values1 = Reference(ws, min_col = 5, min_row = 1, max_col = 5, max_row = len(imeUcenika)+1)
    values2 = Reference(ws, min_col = 6,min_row = 1, max_col = 6, max_row = len(imeUcenika)+1)
    values3 = Reference(ws, min_col = 3, min_row = 1, max_col = 3, max_row = len(imeUcenika)+1)
    x_values = Reference(ws, min_col = 2, min_row = 2, max_col = 2, max_row = len(imeUcenika)+1)

    chart = BarChart()
    chart.add_data(values1, titles_from_data = True)
    chart.add_data(values2, titles_from_data = True)
    # chart.add_data(values3, titles_from_data = True)
    chart.set_categories(x_values)
    chart.title = "Grafikon studenata"
    chart.x_axis.title = "Student"
    chart.y_axis.title = "Ocena"
    chart.y_axis.scaling.min = 5
    chart.y_axis.scaling.max = 10
    chart2 = LineChart()
    chart2.add_data(values3, titles_from_data = True)
    chart2.y_axis.axId = 200
    chart2.y_axis.title = "Godina upisa"
    chart.y_axis.crosses = "max"
    chart +=  chart2
    ws1.add_chart(chart, 'D6')

if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('-g','--graph',help='Creates a graph of already given data.',action='store_true')
    parser.add_argument('-s','--addStudents',help='Calls a function that let you add students to your excel file.',action='store_true')
    parser.add_argument('-a','--average',help='Calls a function that adds new row to excel file which calculates average amount of years studied, average grade and highest grade.',action='store_true')
    args = parser.parse_args()

    graph = args.graph
    students = args.addStudents
    average = args.average

    addStudents()
    if graph:
        createChart()
    if average:
        averageAmount()
    wb.save('studentiFTN.xlsx')